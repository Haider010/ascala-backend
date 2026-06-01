import csv
import io
import json
import mimetypes
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import PurePosixPath
from typing import Any

import requests
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook

from app.db.schema import ensure_installed_locations_table

GHL_MEDIA_UPLOAD_URL = "https://services.leadconnectorhq.com/medias/upload-file"
GHL_LOCATION_TOKEN_URL = "https://services.leadconnectorhq.com/oauth/locationToken"
GHL_API_VERSION = "2023-02-21"
MAX_POST_ROWS = 90
MAX_MEDIA_FILES = 90
MAX_MEDIA_BYTES = 50 * 1024 * 1024

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".heic", ".heif"}
VIDEO_EXTENSIONS = {".mp4", ".mov", ".m4v", ".webm", ".avi", ".mkv"}
GIF_EXTENSIONS = {".gif"}
MEDIA_EXTENSIONS = IMAGE_EXTENSIONS | VIDEO_EXTENSIONS | GIF_EXTENSIONS

DATE_FORMATS = (
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d %H:%M",
    "%Y/%m/%d %H:%M:%S",
    "%m/%d/%Y %H:%M",
    "%m/%d/%Y %H:%M:%S",
    "%m-%d-%Y %H:%M",
    "%m-%d-%Y %H:%M:%S",
)


@dataclass(frozen=True)
class MediaFile:
    name: str
    data: bytes
    content_type: str
    media_kind: str


@dataclass(frozen=True)
class PreparedCsv:
    content: str
    filename: str
    summary: dict[str, Any]


def _location_token_expires_at(expires_in: int | str | None) -> datetime:
    try:
        seconds = int(expires_in or 0)
    except (TypeError, ValueError):
        seconds = 0
    if seconds <= 0:
        seconds = 60 * 60 * 24
    return datetime.now(timezone.utc) + timedelta(seconds=max(seconds - 300, 60))


def _request_location_token(company_access_token: str, company_id: str, location_id: str) -> dict[str, Any]:
    try:
        response = requests.post(
            GHL_LOCATION_TOKEN_URL,
            data={"companyId": company_id, "locationId": location_id},
            headers={
                "Accept": "application/json",
                "Content-Type": "application/x-www-form-urlencoded",
                "Version": GHL_API_VERSION,
                "Authorization": f"Bearer {company_access_token}",
            },
            timeout=30,
        )
    except requests.RequestException as exc:
        raise HTTPException(status_code=502, detail="Unable to request a GHL location access token.") from exc

    if response.status_code >= 400:
        detail = response.text[:500] or response.reason
        raise HTTPException(status_code=502, detail=f"GHL location token request failed: {detail}")

    try:
        payload: dict[str, Any] = response.json()
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=502, detail="GHL location token response was not valid JSON.") from exc

    if not payload.get("access_token"):
        raise HTTPException(status_code=502, detail="GHL location token response did not include an access token.")

    return payload


def _store_location_token(cursor, installed_location_id: str, payload: dict[str, Any]) -> str:
    cursor.execute(
        """
        UPDATE ascala_installed_locations
        SET
            location_access_token = %s,
            location_refresh_token = %s,
            location_token_type = %s,
            location_token_scope = %s,
            location_refresh_token_id = %s,
            location_token_expires_at = %s,
            location_token_updated_at = NOW()
        WHERE id = %s
        """,
        (
            payload.get("access_token"),
            payload.get("refresh_token"),
            payload.get("token_type"),
            payload.get("scope"),
            payload.get("refreshTokenId"),
            _location_token_expires_at(payload.get("expires_in")),
            installed_location_id,
        ),
    )
    return payload["access_token"]


def get_location_access_token(cursor, location_id: str) -> str:
    ensure_installed_locations_table(cursor)
    cursor.execute(
        """
        SELECT
            l.id,
            l.company_id,
            l.location_access_token,
            l.location_token_expires_at,
            c.access_token,
            c.scope
        FROM ascala_installed_locations l
        JOIN ascala_connections c ON c.id = l.connection_id
        WHERE l.location_id = %s
        ORDER BY l.last_seen_at DESC
        LIMIT 1
        """,
        (location_id,),
    )
    row = cursor.fetchone()
    if not row:
        raise HTTPException(status_code=409, detail="This location is not connected to a stored HighLevel install.")

    installed_location_id, company_id, location_token, expires_at, company_token, company_scope = row
    now = datetime.now(timezone.utc)
    if expires_at and expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if location_token and expires_at and expires_at > now + timedelta(minutes=5):
        return location_token

    if not company_token:
        raise HTTPException(status_code=409, detail="GHL company access token is not available for this location.")
    if "oauth.write" not in (company_scope or ""):
        raise HTTPException(
            status_code=409,
            detail="The HighLevel app install is missing oauth.write. Reinstall the app after adding the scope.",
        )

    payload = _request_location_token(company_token, company_id, location_id)
    return _store_location_token(cursor, installed_location_id, payload)


def _clean_upload_filename(filename: str | None) -> str:
    name = PurePosixPath(filename or "social-planner-upload").name
    return re.sub(r"[^a-zA-Z0-9._-]+", "-", name).strip("-") or "social-planner-upload"


def _extension(filename: str) -> str:
    return PurePosixPath(filename).suffix.lower()


def _media_kind(filename: str) -> str:
    extension = _extension(filename)
    if extension in IMAGE_EXTENSIONS:
        return "image"
    if extension in GIF_EXTENSIONS:
        return "gif"
    if extension in VIDEO_EXTENSIONS:
        return "video"
    raise HTTPException(status_code=400, detail=f"Unsupported media file type: {filename}")


def _is_ignored_zip_entry(name: str) -> bool:
    path = PurePosixPath(name)
    return (
        name.endswith("/")
        or any(part == "__MACOSX" for part in path.parts)
        or path.name.startswith(".")
    )


async def read_media_zip(media_zip: UploadFile) -> list[MediaFile]:
    if not media_zip.filename or _extension(media_zip.filename) != ".zip":
        raise HTTPException(status_code=400, detail="Upload a ZIP file containing the media assets.")

    zip_bytes = await media_zip.read()
    try:
        archive = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except zipfile.BadZipFile as exc:
        raise HTTPException(status_code=400, detail="The media ZIP file could not be opened.") from exc

    media_files: list[MediaFile] = []
    total_size = 0
    for info in archive.infolist():
        if _is_ignored_zip_entry(info.filename):
            continue
        filename = PurePosixPath(info.filename).name
        if _extension(filename) not in MEDIA_EXTENSIONS:
            continue
        total_size += info.file_size
        if total_size > MAX_MEDIA_BYTES:
            raise HTTPException(status_code=413, detail="Media ZIP is too large. Keep media files under 50 MB total.")
        if len(media_files) >= MAX_MEDIA_FILES:
            raise HTTPException(status_code=400, detail=f"GHL CSV import supports up to {MAX_MEDIA_FILES} media/post rows.")
        data = archive.read(info)
        content_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
        media_files.append(MediaFile(filename, data, content_type, _media_kind(filename)))

    if not media_files:
        raise HTTPException(status_code=400, detail="No supported media files were found in the ZIP.")

    return media_files


async def read_social_planner_rows(schedule_file: UploadFile) -> tuple[list[list[str]], str]:
    filename = schedule_file.filename or ""
    extension = _extension(filename)
    file_bytes = await schedule_file.read()

    if extension == ".csv":
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = file_bytes.decode("latin-1")
        return list(csv.reader(io.StringIO(text))), "csv"

    if extension == ".xlsx":
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        sheet = workbook.active
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(["" if value is None else str(value) for value in row])
        return rows, "xlsx"

    raise HTTPException(status_code=400, detail="Upload a GHL CSV or XLSX schedule file.")


def _pad_rows(rows: list[list[str]]) -> list[list[str]]:
    width = max((len(row) for row in rows), default=0)
    return [row + [""] * (width - len(row)) for row in rows]


def _header_index(headers: list[str], expected: str) -> int | None:
    try:
        return headers.index(expected)
    except ValueError:
        return None


def _validate_date(value: str, column_name: str, row_number: int) -> str:
    trimmed = value.strip()
    if not trimmed:
        return value

    for date_format in DATE_FORMATS:
        try:
            parsed = datetime.strptime(trimmed, date_format)
            return parsed.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue

    raise HTTPException(
        status_code=400,
        detail=f"Row {row_number} has an unsupported date in '{column_name}'. Use a GHL-supported date format.",
    )


def _normalize_dates(rows: list[list[str]], field_headers: list[str], post_rows: list[list[str]]) -> None:
    date_columns = [
        (index, header)
        for index, header in enumerate(field_headers)
        if "date" in header.lower() or "postatspecifictime" in header.lower()
    ]

    for post_index, row in enumerate(post_rows, start=3):
        for column_index, header in date_columns:
            if column_index < len(row):
                row[column_index] = _validate_date(row[column_index], header, post_index)


def _normalize_booleans(field_headers: list[str], post_rows: list[list[str]]) -> None:
    boolean_columns = [
        index
        for index, header in enumerate(field_headers)
        if "true/false" in header.lower()
    ]

    for row in post_rows:
        for column_index in boolean_columns:
            if column_index >= len(row):
                continue
            value = row[column_index].strip()
            if value.lower() == "true":
                row[column_index] = "TRUE"
            elif value.lower() == "false":
                row[column_index] = "FALSE"


def _extract_urls(payload: Any) -> list[str]:
    urls: list[str] = []
    if isinstance(payload, str):
        urls.extend(re.findall(r"https?://[^\s\"'<>]+", payload))
    elif isinstance(payload, dict):
        for value in payload.values():
            urls.extend(_extract_urls(value))
    elif isinstance(payload, list):
        for value in payload:
            urls.extend(_extract_urls(value))
    return urls


def _pick_media_url(payload: Any) -> str:
    urls = _extract_urls(payload)
    if not urls:
        raise HTTPException(status_code=502, detail="GHL media upload succeeded but did not return a media URL.")

    for url in urls:
        if "storage.googleapis.com" in url or "msgsndr" in url:
            return url.rstrip(".,")
    return urls[0].rstrip(".,")


def upload_media_file(media_file: MediaFile, token: str) -> str:
    headers = {
        "Accept": "application/json",
        "Version": GHL_API_VERSION,
        "Authorization": f"Bearer {token}",
    }
    files = {
        "file": (media_file.name, io.BytesIO(media_file.data), media_file.content_type),
    }
    try:
        response = requests.post(GHL_MEDIA_UPLOAD_URL, headers=headers, files=files, timeout=120)
    except requests.RequestException as exc:
        raise HTTPException(status_code=502, detail=f"GHL media upload failed for {media_file.name}.") from exc

    if response.status_code >= 400:
        detail = response.text[:500] or response.reason
        raise HTTPException(status_code=502, detail=f"GHL rejected {media_file.name}: {detail}")

    try:
        payload: Any = response.json()
    except json.JSONDecodeError:
        payload = response.text

    return _pick_media_url(payload)


def _set_media_url(row: list[str], field_headers: list[str], media_file: MediaFile, url: str, row_number: int) -> str:
    image_index = _header_index(field_headers, "imageUrls (comma-separated)")
    gif_index = _header_index(field_headers, "gifUrl")
    video_index = _header_index(field_headers, "videoUrls (comma-separated)")

    if media_file.media_kind == "image" and image_index is not None:
        row[image_index] = url
        return "imageUrls"
    if media_file.media_kind == "gif" and gif_index is not None:
        row[gif_index] = url
        return "gifUrl"
    if media_file.media_kind == "video" and video_index is not None:
        row[video_index] = url
        return "videoUrls"

    raise HTTPException(
        status_code=400,
        detail=f"Row {row_number} cannot receive {media_file.media_kind} media because the matching GHL column is missing.",
    )


def _to_csv(rows: list[list[str]]) -> str:
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerows(rows)
    return output.getvalue()


async def build_uply_csv(schedule_file: UploadFile, media_zip: UploadFile, token: str | None) -> PreparedCsv:
    if not token:
        raise HTTPException(status_code=500, detail="GHL media upload token is not configured.")

    rows, source_type = await read_social_planner_rows(schedule_file)
    rows = _pad_rows(rows)
    if len(rows) < 3:
        raise HTTPException(status_code=400, detail="The GHL schedule file must include two header rows and at least one post row.")

    field_headers = rows[1]
    post_rows = [row for row in rows[2:] if any(cell.strip() for cell in row)]
    if len(post_rows) > MAX_POST_ROWS:
        raise HTTPException(status_code=400, detail=f"CSV import supports up to {MAX_POST_ROWS} posts per file.")

    media_files = await read_media_zip(media_zip)
    if len(media_files) > len(post_rows):
        raise HTTPException(status_code=400, detail="The ZIP contains more media files than post rows in the schedule.")

    _normalize_dates(rows, field_headers, post_rows)
    _normalize_booleans(field_headers, post_rows)

    uploaded: list[dict[str, str]] = []
    for index, media_file in enumerate(media_files):
        post_row = post_rows[index]
        row_number = index + 3
        media_url = upload_media_file(media_file, token)
        column = _set_media_url(post_row, field_headers, media_file, media_url, row_number)
        uploaded.append({
            "fileName": media_file.name,
            "mediaKind": media_file.media_kind,
            "column": column,
            "url": media_url,
        })

    base_name = _clean_upload_filename(schedule_file.filename)
    base_name = re.sub(r"\.(csv|xlsx)$", "", base_name, flags=re.IGNORECASE)
    filename = f"uply-ghl-ready-{base_name}.csv"
    summary = {
        "sourceType": source_type,
        "postRows": len(post_rows),
        "mediaUploaded": len(uploaded),
        "unmatchedPostRows": max(len(post_rows) - len(media_files), 0),
        "outputFilename": filename,
        "uploads": uploaded,
    }

    return PreparedCsv(content=_to_csv(rows), filename=filename, summary=summary)
